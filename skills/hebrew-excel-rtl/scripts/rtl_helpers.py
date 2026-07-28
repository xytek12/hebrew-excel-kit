# -*- coding: utf-8 -*-
"""Helpers for building Hebrew RTL Excel workbooks with openpyxl.

Import these instead of re-deriving the styling each time:

    from rtl_helpers import (
        FMT, rtl_workbook, style_header, write_row, autofit, add_table, banner
    )
"""
from __future__ import annotations

import datetime
import os
from typing import Any, Iterable, Sequence

from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.worksheet.worksheet import Worksheet

# --------------------------------------------------------------------------
# Number formats. Every one of these was rendered through real Excel (Office 16,
# Hebrew locale) and produces the output shown in the comment.
# --------------------------------------------------------------------------
FMT = {
    "shekel":       r'#,##0.00\ [$₪-40D]',                                          # 39.90 ₪
    "shekel_whole": r'#,##0\ [$₪-40D]',                                             # 1,250 ₪
    "shekel_neg":   r'#,##0.00\ [$₪-40D];[Red]\(#,##0.00\ [$₪-40D]\)',              # (2,396.00 ₪)
    "shekel_dash":  r'#,##0.00\ [$₪-40D];[Red]\(#,##0.00\ [$₪-40D]\);"—"',          # —
    "thousands":    r'#,##0,\ "אלפי ₪"',                                            # 1,250 אלפי ₪
    "millions":     r'#,##0.0,,\ "מ׳ ₪"',                                           # 1.3 מ׳ ₪
    "date":         "DD/MM/YYYY",                                                   # 15/01/2026
    "date_long":    r'D\ בMMMM\ YYYY',                                              # 15 בינואר 2026
    "month_year":   r'MMMM\ YYYY',                                                  # ינואר 2026
    "time":         "HH:MM",                                                        # 14:30
    "int":          "#,##0",                                                        # 1,250
    "dec2":         "#,##0.00",                                                     # 1,250.00
    "pct":          "0.0%",                                                         # 12.5%
    "pct_signed":   "+0.0%;-0.0%;0.0%",                                             # +12.5%
    "variance":     '+0.0%;[Red]-0.0%;"—"',                                         # -3.2% (red)
    "qty":          r'#,##0\ "יח׳"',                                                # 120 יח׳
}

# Arial is the only font guaranteed to carry Hebrew on every Office install,
# Windows and Mac alike. Do not swap it for a Google font in a file you email.
FONT_NAME = "Arial"

NAVY = "1F3864"
LIGHT = "D9E2F3"
GREY = "F2F2F2"

_thin = Side(style="thin", color="BFBFBF")
BORDER = Border(left=_thin, right=_thin, top=_thin, bottom=_thin)


def rtl_alignment(horizontal: str = "right", **kw: Any) -> Alignment:
    """RTL alignment. readingOrder=2 is what fixes bidi *inside* the cell."""
    kw.setdefault("vertical", "center")
    return Alignment(horizontal=horizontal, readingOrder=2, **kw)


RTL = rtl_alignment()
RTL_CENTER = rtl_alignment("center")


def rtl_workbook(wb) -> None:
    """Flip every sheet in the workbook to right-to-left.

    This is per *sheet*, not per workbook. Call it once at the end of the build,
    after all sheets exist, so no sheet is missed.
    """
    for ws in wb.worksheets:
        ws.sheet_view.rightToLeft = True


def style_header(ws: Worksheet, row: int, headers: Sequence[str],
                 start_col: int = 1, fill: str = NAVY) -> None:
    """Write and style a header row."""
    for i, text in enumerate(headers):
        c = ws.cell(row=row, column=start_col + i, value=text)
        c.font = Font(name=FONT_NAME, bold=True, color="FFFFFF", size=11)
        c.fill = PatternFill("solid", fgColor=fill)
        c.alignment = rtl_alignment("center", wrap_text=True)
        c.border = BORDER


def write_row(ws: Worksheet, row: int, values: Sequence[Any],
              formats: Sequence[str | None] | None = None,
              start_col: int = 1, bold: bool = False,
              fill: str | None = None) -> None:
    """Write one data row with RTL alignment and per-column number formats.

    `formats` is a sequence of FMT keys (or raw format codes) aligned to `values`;
    use None to leave a cell's format alone.
    """
    for i, v in enumerate(values):
        c = ws.cell(row=row, column=start_col + i, value=v)
        c.alignment = RTL
        c.font = Font(name=FONT_NAME, bold=bold, size=11)
        c.border = BORDER
        if fill:
            c.fill = PatternFill("solid", fgColor=fill)
        if formats and i < len(formats) and formats[i]:
            c.number_format = FMT.get(formats[i], formats[i])


def banner(ws: Worksheet, row: int, text: str, span: int, start_col: int = 1) -> None:
    """A centred title across `span` columns WITHOUT merging.

    Merged cells break sorting, filtering and pivots. centerContinuous gives the
    same look and keeps the range clean. Note the camelCase — `center_continuous`
    raises ValueError in openpyxl.
    """
    c = ws.cell(row=row, column=start_col, value=text)
    c.font = Font(name=FONT_NAME, bold=True, size=14, color=NAVY)
    for i in range(span):
        ws.cell(row=row, column=start_col + i).alignment = Alignment(
            horizontal="centerContinuous", vertical="center", readingOrder=2
        )


def add_table(ws: Worksheet, name: str, ref: str,
              style: str = "TableStyleMedium9") -> None:
    """Add an Excel table. `name` must be unique and contain no spaces."""
    if " " in name:
        raise ValueError(f"table name must not contain spaces: {name!r}")
    t = Table(displayName=name, ref=ref)
    t.tableStyleInfo = TableStyleInfo(
        name=style, showRowStripes=True, showColumnStripes=False
    )
    ws.add_table(t)


def autofit(ws: Worksheet, min_w: int = 12, max_w: int = 50,
            factor: float = 1.3) -> None:
    """Estimate column widths.

    There is no auto-fit in the xlsx format — Excel computes it at render time and
    openpyxl cannot measure text. Hebrew renders wider than the same character
    count in Latin, hence the 1.3 factor; 1.0 gives visibly cramped columns.
    """
    for col in ws.columns:
        longest = 0
        letter = None
        for cell in col:
            if letter is None:
                letter = cell.column_letter
            v = cell.value
            if v is None:
                continue
            if isinstance(v, (datetime.date, datetime.datetime)):
                n = 10
            else:
                n = len(str(v))
            longest = max(longest, n)
        if letter:
            ws.column_dimensions[letter].width = max(
                min_w, min(max_w, int(longest * factor) + 4)
            )


def freeze_header(ws: Worksheet, first_data_row: int = 2) -> None:
    """Freeze above the first data row. Excel mirrors this correctly on RTL sheets."""
    ws.freeze_panes = f"A{first_data_row}"


def suppress_text_number_warning(path: str, ranges: dict[str, str]) -> None:
    """Silence Excel's green "number stored as text" triangle on given ranges.

    Call this **after** `wb.save(path)` — it rewrites the saved file.

        wb.save(OUT)
        suppress_text_number_warning(OUT, {"מלאי": "D4:D15"})

    Why a post-save pass: openpyxl 3.1.5 ships an `IgnoredErrors` class but never
    writes it — `Worksheet` has no `ignored_errors` attribute, so assigning one is a
    silent no-op. The `<ignoredErrors>` element has to be injected into the sheet XML
    directly.

    Some columns are legitimately text that merely looks numeric: clothing sizes
    (32, L), SKUs, invoice numbers, ID numbers that must keep leading zeros, phone
    numbers. Converting them to real numbers corrupts them. A grid dotted with green
    triangles reads as broken, so suppress the warning instead.
    """
    import re
    import shutil
    import tempfile
    import zipfile
    from xml.sax.saxutils import quoteattr

    NS_MAIN = "{http://schemas.openxmlformats.org/spreadsheetml/2006/main}"
    NS_REL = "{http://schemas.openxmlformats.org/officeDocument/2006/relationships}"
    import xml.etree.ElementTree as ET

    src = zipfile.ZipFile(path, "r")
    wb_xml = ET.fromstring(src.read("xl/workbook.xml"))
    rels = ET.fromstring(src.read("xl/_rels/workbook.xml.rels"))
    rel_target = {
        r.get("Id"): r.get("Target")
        for r in rels
    }

    # sheet display name -> xl/worksheets/sheetN.xml
    sheet_part: dict[str, str] = {}
    for sh in wb_xml.iter(f"{NS_MAIN}sheet"):
        rid = sh.get(f"{NS_REL}id")
        target = rel_target.get(rid, "")
        if target.startswith("/xl/"):
            part = target[1:]
        elif target.startswith("xl/"):
            part = target
        else:
            part = "xl/" + target.lstrip("/")
        sheet_part[sh.get("name")] = part

    unknown = set(ranges) - set(sheet_part)
    if unknown:
        raise KeyError(f"sheet(s) not in workbook: {sorted(unknown)}")

    # elements that must come AFTER <ignoredErrors> per the CT_Worksheet sequence
    AFTER = ("<smartTags", "<drawing", "<legacyDrawing", "<picture",
             "<oleObjects", "<controls", "<tableParts", "<extLst")

    patches: dict[str, bytes] = {}
    for sheet_name, sqref in ranges.items():
        part = sheet_part[sheet_name]
        xml = src.read(part).decode("utf-8")
        if "<ignoredErrors" in xml:
            continue
        frag = (f"<ignoredErrors><ignoredError sqref={quoteattr(sqref)} "
                f'numberStoredAsText="1"/></ignoredErrors>')
        pos = len(xml)
        for tag in AFTER:
            m = re.search(re.escape(tag) + r"[ />]", xml)
            if m and m.start() < pos:
                pos = m.start()
        if pos == len(xml):
            pos = xml.rindex("</worksheet>")
        patches[part] = (xml[:pos] + frag + xml[pos:]).encode("utf-8")

    tmp_fd, tmp = tempfile.mkstemp(suffix=".xlsx")
    os.close(tmp_fd)
    with zipfile.ZipFile(tmp, "w", zipfile.ZIP_DEFLATED) as out:
        for item in src.infolist():
            data = patches.get(item.filename) or src.read(item.filename)
            out.writestr(item, data)
    src.close()
    shutil.move(tmp, path)


def sheet_ref(sheet_name: str, addr: str) -> str:
    """Build a cross-sheet formula reference, quoting Hebrew sheet names."""
    if any(ch in sheet_name for ch in " '\"-"):
        return f"'{sheet_name}'!{addr}"
    return f"{sheet_name}!{addr}"

# -*- coding: utf-8 -*-
"""A visual system for Hebrew Excel reports that non-financial managers can read.

The default openpyxl look — grey fills, borders on every cell, a different colour per
concept — is exhausting to read. So is the financial-modelling convention of colouring
inputs blue and formulas black: it is a *modeller's* debugging aid, and to a manager it
just looks like the file has two kinds of money in it.

This module encodes one restrained system:

  * one accent colour, used only for headers and titles
  * data on white, with an optional near-invisible zebra tint
  * horizontal rules only — no vertical grid lines inside data
  * colour on a number means exactly one thing: something needs attention
  * every number in a column is the same colour and the same format

Colour tokens are adapted from the `israeli-ui-design-system` skill's palette so a
workbook, a dashboard, and a web UI for the same company look related.
"""
from __future__ import annotations

from typing import Any, Sequence

from openpyxl.formatting.rule import CellIsRule, FormulaRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

# --------------------------------------------------------------------------
# Palette
# --------------------------------------------------------------------------
NAVY = "1F3A5C"        # headers, titles
NAVY_SOFT = "E8EDF4"   # section bands, KPI tiles
ZEBRA = "F7F9FC"       # alternating data rows — barely visible on purpose
RULE = "D8DEE8"        # horizontal rules
INK = "1A1A1A"         # all body numbers and text
MUTED = "6B7280"       # captions, notes, units

GOOD = "16A34A"
WARN = "D97706"
BAD = "DC2626"
INFO = "2563EB"
BAD_BG = "FDE8E8"
WARN_BG = "FEF3E2"
GOOD_BG = "E9F7EF"
INFO_BG = "E8F0FE"

# Arial is the only font with guaranteed Hebrew coverage on every Office install,
# Windows and Mac. Heebo/Rubik/Assistant look better but must be installed on the
# reader's machine — never use them in a workbook you email.
FONT = "Arial"

DEFAULT_COL_WIDTH = 13.0

_rule = Side(style="thin", color=RULE)
UNDERLINE = Border(bottom=_rule)
NO_BORDER = Border()


def _al(h: str = "right", **kw: Any) -> Alignment:
    kw.setdefault("vertical", "center")
    return Alignment(horizontal=h, readingOrder=2, **kw)


# --------------------------------------------------------------------------
# Structure
# --------------------------------------------------------------------------
def page_title(ws: Worksheet, row: int, text: str, subtitle: str | None = None) -> int:
    """Big page title, no fill. Returns the next free row."""
    c = ws.cell(row=row, column=1, value=text)
    c.font = Font(name=FONT, size=18, bold=True, color=NAVY)
    c.alignment = _al()
    ws.row_dimensions[row].height = 30
    if subtitle:
        s = ws.cell(row=row + 1, column=1, value=subtitle)
        s.font = Font(name=FONT, size=10, color=MUTED)
        s.alignment = _al()
        ws.row_dimensions[row + 1].height = 16
        return row + 3
    return row + 2


def section(ws: Worksheet, row: int, text: str, span: int) -> int:
    """A section band. Returns the next free row."""
    for i in range(span):
        cell = ws.cell(row=row, column=1 + i)
        cell.fill = PatternFill("solid", fgColor=NAVY_SOFT)
        cell.alignment = _al()
    c = ws.cell(row=row, column=1, value=text)
    c.font = Font(name=FONT, size=12, bold=True, color=NAVY)
    ws.row_dimensions[row].height = 24
    return row + 1


def header_row(ws: Worksheet, row: int, headers: Sequence[str],
               widths: Sequence[int] | None = None) -> int:
    """The one dark band in the sheet."""
    for i, text in enumerate(headers):
        c = ws.cell(row=row, column=1 + i, value=text)
        c.font = Font(name=FONT, size=11, bold=True, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor=NAVY)
        c.alignment = _al("center", wrap_text=True)
    if widths:
        for i, w in enumerate(widths):
            ws.column_dimensions[get_column_letter(1 + i)].width = w

    # Headers wrap, and Excel does not auto-grow a row whose height you have set. Size the
    # row from the longest header against its own column width, or two-line headings like
    # "שבועות כיסוי ממוצע" get their second line clipped.
    lines = 1
    for i, text in enumerate(headers):
        w = widths[i] if widths and i < len(widths) else DEFAULT_COL_WIDTH
        lines = max(lines, -(-len(str(text)) // max(6, int(w * 0.75))))
    ws.row_dimensions[row].height = 20 + 14 * min(lines, 3)
    return row + 1


def data_row(ws: Worksheet, row: int, values: Sequence[Any],
             formats: Sequence[str | None] | None = None,
             bold: bool = False, align: Sequence[str] | None = None) -> None:
    """One data row. White background, single ink colour, rule underneath."""
    for i, v in enumerate(values):
        c = ws.cell(row=row, column=1 + i, value=v)
        c.font = Font(name=FONT, size=11, bold=bold, color=INK)
        c.alignment = _al(align[i] if align and i < len(align) else "right")
        c.border = UNDERLINE
        if formats and i < len(formats) and formats[i]:
            c.number_format = formats[i]
    ws.row_dimensions[row].height = 20


def total_row(ws: Worksheet, row: int, values: Sequence[Any],
              formats: Sequence[str | None] | None = None) -> None:
    """A totals line — bold on a soft band, not another dark bar."""
    for i, v in enumerate(values):
        c = ws.cell(row=row, column=1 + i, value=v)
        c.font = Font(name=FONT, size=11, bold=True, color=NAVY)
        c.fill = PatternFill("solid", fgColor=NAVY_SOFT)
        c.alignment = _al()
        c.border = Border(top=Side(style="thin", color=NAVY))
        if formats and i < len(formats) and formats[i]:
            c.number_format = formats[i]
    ws.row_dimensions[row].height = 22


def zebra(ws: Worksheet, first_row: int, last_row: int, last_col: int) -> None:
    """Tint every other row. Keep it subtle enough that you notice it only if it's gone."""
    for r in range(first_row, last_row + 1):
        if (r - first_row) % 2 == 1:
            for c in range(1, last_col + 1):
                cell = ws.cell(row=r, column=c)
                if cell.fill.fgColor.rgb in (None, "00000000"):
                    cell.fill = PatternFill("solid", fgColor=ZEBRA)


# --------------------------------------------------------------------------
# Dashboard pieces
# --------------------------------------------------------------------------
def kpi(ws: Worksheet, row: int, col: int, label: str, value: Any,
        number_format: str, note: str | None = None, width: int = 2) -> None:
    """A KPI tile: small grey label, large navy number, optional caption underneath.

    Occupies 3 rows × `width` columns starting at (row, col).

    The text is centred with `centerContinuous` across the tile rather than by merging
    cells. Same look, and it keeps the no-merge rule intact — a big number in a single
    15-wide column renders as ######, but centerContinuous lets it spill across the
    whole tile.
    """
    for r in range(row, row + 3):
        for c in range(col, col + width):
            cell = ws.cell(row=r, column=c)
            cell.fill = PatternFill("solid", fgColor=NAVY_SOFT)
            cell.alignment = _al("centerContinuous")

    lab = ws.cell(row=row, column=col, value=label)
    lab.font = Font(name=FONT, size=10, bold=True, color=MUTED)
    lab.alignment = _al("centerContinuous")
    ws.row_dimensions[row].height = 18

    val = ws.cell(row=row + 1, column=col, value=value)
    val.font = Font(name=FONT, size=16, bold=True, color=NAVY)
    val.number_format = number_format
    val.alignment = _al("centerContinuous")
    ws.row_dimensions[row + 1].height = 28

    cap = ws.cell(row=row + 2, column=col, value=note or "")
    cap.font = Font(name=FONT, size=9, color=MUTED)
    cap.alignment = _al("centerContinuous")
    ws.row_dimensions[row + 2].height = 16


def insight(ws: Worksheet, row: int, formula: str, span: int) -> None:
    """One plain-Hebrew sentence, computed live from the data.

    This is what turns a grid of numbers into something a manager reads. Build the
    text with a formula so it can never drift out of date:

        '="החנות המובילה: "&INDEX(...)&" עם "&TEXT(...,"#,##0")&" ש\\"ח"'
    """
    c = ws.cell(row=row, column=1, value=formula)
    c.font = Font(name=FONT, size=11, color=INK)
    c.alignment = _al(wrap_text=False)
    for i in range(1, span):
        ws.cell(row=row, column=1 + i).alignment = _al()
    ws.row_dimensions[row].height = 20


def status_colours(ws: Worksheet, ref: str,
                   bad: Sequence[str] = (), warn: Sequence[str] = (),
                   good: Sequence[str] = (), info: Sequence[str] = ()) -> None:
    """Colour a text status column by its value.

    Four buckets, not three, because "about to run out" and "far too much stock" are
    opposite problems and must not share a colour — a manager scanning the column will
    otherwise read overstock as urgency. `info` is the muted blue bucket: worth knowing,
    nothing to do today.

    The cell keeps its text. Colour supports the word, it does not replace it — nobody
    should have to decode a legend to read a status.
    """
    for words, fg, bg in ((bad, BAD, BAD_BG), (warn, WARN, WARN_BG),
                          (good, GOOD, GOOD_BG), (info, INFO, INFO_BG)):
        for w in words:
            ws.conditional_formatting.add(
                ref,
                CellIsRule(operator="equal", formula=[f'"{w}"'],
                           font=Font(name=FONT, size=11, bold=True, color=fg),
                           fill=PatternFill("solid", fgColor=bg)),
            )


def flag_negative(ws: Worksheet, ref: str) -> None:
    """Red only where a number is genuinely a problem."""
    ws.conditional_formatting.add(
        ref,
        CellIsRule(operator="lessThan", formula=["0"],
                   font=Font(name=FONT, size=11, bold=True, color=BAD)),
    )


def flag_below(ws: Worksheet, ref: str, threshold_formula: str) -> None:
    """Red where a value falls below a per-row threshold (e.g. stock < reorder point)."""
    ws.conditional_formatting.add(
        ref,
        FormulaRule(formula=[threshold_formula],
                    font=Font(name=FONT, size=11, bold=True, color=BAD),
                    fill=PatternFill("solid", fgColor=BAD_BG)),
    )


def note(ws: Worksheet, row: int, text: str) -> int:
    c = ws.cell(row=row, column=1, value=text)
    c.font = Font(name=FONT, size=9, italic=True, color=MUTED)
    c.alignment = _al()
    return row + 1


def hide_gridlines(ws: Worksheet) -> None:
    """Excel's default gridlines fight with your own rules. Turn them off on any
    sheet that has real formatting — especially dashboards."""
    ws.sheet_view.showGridLines = False

# -*- coding: utf-8 -*-
"""Build the Hebrew RTL demo workbook.

Fictional company: אופנת גלים בע״מ — an Israeli clothing chain with 6 branches.
All figures are invented.

Design goal: a manager with no finance background should understand every sheet
without anyone explaining it. See excel_design.py for the visual rules.

    py build_demo.py
"""
from __future__ import annotations

import datetime
import os
import sys

import openpyxl
from openpyxl.chart import BarChart, LineChart, Reference
from openpyxl.chart.label import DataLabelList
from openpyxl.utils import get_column_letter

HERE = os.path.dirname(os.path.abspath(__file__))
sys.path.insert(0, os.path.join(HERE, "..", "skills", "hebrew-excel-rtl", "scripts"))

from excel_design import (  # noqa: E402
    BAD, GOOD, INK, MUTED, NAVY, NAVY_SOFT, WARN,
    data_row, flag_negative, header_row, hide_gridlines, insight, kpi, note,
    page_title, section, status_colours, total_row, zebra,
)
from rtl_helpers import (  # noqa: E402
    FMT, add_table, freeze_header, rtl_workbook, suppress_text_number_warning,
)

OUT = os.path.join(HERE, "דוח-כספי-לדוגמה.xlsx")

MONTHS = ["ינואר", "פברואר", "מרץ", "אפריל", "מאי", "יוני",
          "יולי", "אוגוסט", "ספטמבר", "אוקטובר", "נובמבר", "דצמבר"]

STORES = [
    "גרנד קניון חיפה",
    "קניון הקריון",
    "חיפה — הדר",
    "עזריאלי תל אביב",
    "קניון מלחה ירושלים",
    "מחסן אונליין",
]
STORE_SIZE = [1.35, 1.15, 0.70, 1.45, 1.00, 0.85]

# Monthly unit sales are per "unit store size" (see STORE_SIZE) and are scaled so the
# chain lands near ₪10.5M gross a year — a plausible size for six branches.
# `cover` is the *target* weeks of stock, chosen to give a realistic spread of statuses
# across the chain rather than an all-green or all-red sheet.
#            מק״ט,     תיאור,                קטגוריה,  מידה,  צבע,     מכר/חודש, כיסוי, מחיר,  עלות
ITEMS = [
    ("TS-1041", "חולצת טי כותנה",      "חולצות",  "M",    "לבן",   119,  6.0, 79.90, 24.90),
    ("TS-1042", "חולצת טי כותנה",      "חולצות",  "L",    "שחור",  146,  1.4, 79.90, 24.90),
    ("SH-2210", "חולצה מכופתרת פשתן",  "חולצות",  "L",    "תכלת",   49,  9.0, 189.00, 68.00),
    ("JN-3305", "מכנסי ג׳ינס סלים",    "מכנסיים", "32",   "כחול",   59,  0.7, 249.00, 96.50),
    ("JN-3307", "מכנסי ג׳ינס ישר",     "מכנסיים", "34",   "שחור",   54, 14.0, 249.00, 96.50),
    ("DR-4102", "שמלת קיץ פרחונית",    "שמלות",   "S",    "ורוד",   38,  3.0, 279.00, 112.00),
    ("DR-4108", "שמלת ערב",            "שמלות",   "M",    "שחור",   16, 16.0, 590.00, 245.00),
    ("JK-5501", "מעיל חורף מרופד",     "מעילים",  "XL",   "אפור",   22,  5.0, 449.00, 189.00),
    ("JK-5504", "ג׳קט ג׳ינס",          "מעילים",  "M",    "כחול",   32,  2.5, 379.00, 154.00),
    ("AC-6001", "חגורת עור",           "אביזרים", "יחיד", "חום",    38,  8.0, 119.00, 42.00),
    ("AC-6014", "צעיף צמר",            "אביזרים", "יחיד", "בורדו",  16, 18.0, 99.00, 38.50),
    ("SK-7203", "חצאית מידי",          "חצאיות",  "S",    "ירוק",   27,  1.8, 199.00, 78.00),
]

SEASONALITY = [1.00, 0.88, 1.05, 0.97, 1.12, 1.21,
               1.18, 1.02, 1.09, 1.14, 1.25, 1.40]

EXPENSES = [
    ("שכר ונלוות",         168_000),
    ("שכר דירה וארנונה",    62_000),
    ("שיווק ופרסום",        41_500),
    ("חשמל, מים ותקשורת",   11_800),
    ("הוצאות משרד",          7_400),
    ("עמלות סליקה",         14_200),
    ("הובלה ושילוח",        19_600),
    ("ביטוח",                8_300),
    ("פחת והפחתות",         12_000),
]


def wobble(i: int, j: int) -> float:
    """Deterministic per-(item, store) variation in demand. No RNG — reproducible."""
    return 0.72 + ((i * 7 + j * 13) % 11) / 14.0


def cover_wobble(i: int, j: int) -> float:
    """Separate variation for stock level, so weeks-of-cover differs between branches.

    Must be independent of `wobble` — if the same factor scaled both units and stock it
    would cancel in the cover ratio and every branch would look identical.
    """
    return 0.70 + ((i * 5 + j * 11) % 13) / 15.0


wb = openpyxl.Workbook()

# ===========================================================================
# הנחות ופרמטרים
# ===========================================================================
ws = wb.active
ws.title = "הנחות ופרמטרים"
hide_gridlines(ws)
r = page_title(ws, 1, "הנחות ופרמטרים",
               "כל המספרים שאפשר לשנות נמצאים כאן. שאר הגיליונות מחושבים מהם אוטומטית.")
r = header_row(ws, r, ["פרמטר", "ערך", "יחידה", "למה זה משפיע"],
               widths=[30, 14, 12, 52])
FIRST_PARAM = r
params = [
    ("שיעור מע״מ",           0.18,  "אחוז", "מופחת מההכנסות כדי להגיע להכנסה נטו"),
    ("שיעור עלות המכר",      0.42,  "אחוז", "כמה עולה לנו הסחורה, כאחוז מההכנסה נטו"),
    ("ימי אשראי לקוחות",     45,    "ימים", "כמה זמן לוקח לגבות מלקוחות"),
    ("ימי אשראי ספקים",      60,    "ימים", "כמה זמן יש לנו לשלם לספקים"),
    ("זמן אספקה מספק",       21,    "ימים", "כמה זמן לוקח לקבל סחורה אחרי הזמנה"),
    ("שבועות כיסוי מינימום",  4,    "שבועות", "מתחת לזה הפריט מסומן כנמוך"),
    ("יתרת פתיחה במזומן",    285_000, "₪",  "כמה מזומן היה בקופה בתחילת השנה"),
]
for i, (name, val, unit, why) in enumerate(params):
    fmt = FMT["pct"] if unit == "אחוז" else (FMT["shekel_whole"] if unit == "₪" else FMT["int"])
    data_row(ws, r + i, [name, val, unit, why], [None, fmt, None, None],
             align=["right", "center", "center", "right"])
r += len(params)
zebra(ws, FIRST_PARAM, r - 1, 4)

VAT = f"'הנחות ופרמטרים'!$B${FIRST_PARAM}"
COGS = f"'הנחות ופרמטרים'!$B${FIRST_PARAM + 1}"
MIN_COVER = f"'הנחות ופרמטרים'!$B${FIRST_PARAM + 5}"
OPENING_CASH = f"'הנחות ופרמטרים'!$B${FIRST_PARAM + 6}"

r += 1
r = section(ws, r, "עונתיות — כמה כל חודש שונה מינואר", 13)
header_row(ws, r, ["חודש"] + MONTHS)
# The params table above already sized A–D; give the remaining month columns a width too,
# or they fall back to the default and clip the Hebrew month names.
for c in range(5, 14):
    ws.column_dimensions[get_column_letter(c)].width = 11
SEASON_ROW = r + 1
data_row(ws, SEASON_ROW, ["מקדם"] + SEASONALITY, [None] + ["0.00"] * 12)
SEASON = f"'הנחות ופרמטרים'!$B${SEASON_ROW}:$M${SEASON_ROW}"
note(ws, SEASON_ROW + 2, "1.00 = כמו ינואר. 1.40 = 40% יותר מינואר.")
freeze_header(ws, FIRST_PARAM)

# ===========================================================================
# מלאי לפי חנות
# ===========================================================================
ws = wb.create_sheet("מלאי לפי חנות")
hide_gridlines(ws)
r = page_title(ws, 1, "מלאי לפי חנות",
               "כל שורה = פריט אחד בחנות אחת. עמודת הסטטוס אומרת מה לעשות.")
INV_HDR = r
r = header_row(ws, r, ["מק״ט", "פריט", "מידה", "קטגוריה", "חנות", "מלאי (יח׳)",
                       "מכירות בחודש (יח׳)", "שבועות כיסוי", "סטטוס",
                       "מחיר מכירה", "עלות ליחידה", "מכירות בחודש (₪)", "שווי מלאי (₪)",
                       "מפתח מיון"],
               widths=[11, 24, 8, 12, 22, 12, 15, 13, 13, 13, 13, 17, 15, 12])
INV_FIRST = r

for j, store in enumerate(STORES):
    for i, (sku, desc, cat, size, colour, units, cover, price, cost) in enumerate(ITEMS):
        u = max(1, round(units * STORE_SIZE[j] * wobble(i, j)))
        s = max(0, round(u / 4.33 * cover * cover_wobble(i, j)))
        data_row(ws, r, [
            sku, desc, size, cat, store, s, u,
            f"=ROUND(F{r}/(G{r}/4.33),1)",
            f'=IF(H{r}<1,"אזל",IF(H{r}<2,"קריטי",IF(H{r}<{MIN_COVER},"נמוך",'
            f'IF(H{r}<=12,"תקין","עודף"))))',
            price, cost, f"=G{r}*J{r}", f"=F{r}*K{r}",
            # Unique sort key. SMALL()/MATCH() on the raw cover column returns the SAME
            # row for every tie, so the alert list would repeat one item. Adding a
            # row-derived epsilon makes every value distinct without moving the ranking.
            f"=H{r}+ROW()/1000000",
        ], [None, None, None, None, None, FMT["int"], FMT["int"], "0.0", None,
            FMT["shekel"], FMT["shekel"], FMT["shekel_whole"], FMT["shekel_whole"],
            "0.000000"],
            align=["center", "right", "center", "right", "right", "center", "center",
                   "center", "center", "right", "right", "right", "right", "center"])
        r += 1

INV_LAST = r - 1
zebra(ws, INV_FIRST, INV_LAST, 13)
add_table(ws, "טבלת_מלאי", f"A{INV_HDR}:N{INV_LAST}")
ws.column_dimensions["N"].hidden = True
status_colours(ws, f"I{INV_FIRST}:I{INV_LAST}",
               bad=["אזל", "קריטי"], warn=["נמוך"], good=["תקין"], info=["עודף"])
freeze_header(ws, INV_FIRST)

INV = "'מלאי לפי חנות'!"
SKU_RNG = f"{INV}$A${INV_FIRST}:$A${INV_LAST}"
ITEM_RNG = f"{INV}$B${INV_FIRST}:$B${INV_LAST}"
SIZE_RNG = f"{INV}$C${INV_FIRST}:$C${INV_LAST}"
STORE_RNG = f"{INV}$E${INV_FIRST}:$E${INV_LAST}"
STOCK_RNG = f"{INV}$F${INV_FIRST}:$F${INV_LAST}"
COVER_RNG = f"{INV}$H${INV_FIRST}:$H${INV_LAST}"
STATUS_RNG = f"{INV}$I${INV_FIRST}:$I${INV_LAST}"
SALES_RNG = f"{INV}$L${INV_FIRST}:$L${INV_LAST}"
VALUE_RNG = f"{INV}$M${INV_FIRST}:$M${INV_LAST}"
KEY_RNG = f"{INV}$N${INV_FIRST}:$N${INV_LAST}"

# ===========================================================================
# סיכום חנויות
# ===========================================================================
ws = wb.create_sheet("סיכום חנויות")
hide_gridlines(ws)
r = page_title(ws, 1, "סיכום חנויות",
               "השוואה בין הסניפים. ככל שהדירוג נמוך יותר, החנות מוכרת יותר.")
STO_HDR = r
r = header_row(ws, r, ["חנות", "מכירות בחודש (₪)", "שווי מלאי (₪)",
                       "פריטים שאזלו", "פריטים קריטיים", "שבועות כיסוי ממוצע",
                       "סבב מלאי חודשי", "דירוג מכירות"],
               widths=[24, 18, 16, 13, 14, 17, 15, 13])
STO_FIRST = r
for store in STORES:
    data_row(ws, r, [
        store,
        f'=SUMIF({STORE_RNG},$A{r},{SALES_RNG})',
        f'=SUMIF({STORE_RNG},$A{r},{VALUE_RNG})',
        f'=COUNTIFS({STORE_RNG},$A{r},{STATUS_RNG},"אזל")',
        f'=COUNTIFS({STORE_RNG},$A{r},{STATUS_RNG},"קריטי")',
        f'=ROUND(AVERAGEIF({STORE_RNG},$A{r},{COVER_RNG}),1)',
        f'=IFERROR(ROUND(B{r}/C{r},2),0)',
        f'=RANK(B{r},$B${STO_FIRST}:$B${STO_FIRST + len(STORES) - 1},0)',
    ], [None, FMT["shekel_whole"], FMT["shekel_whole"], FMT["int"], FMT["int"],
        "0.0", "0.00", FMT["int"]],
        align=["right"] + ["center"] * 7)
    r += 1
STO_LAST = r - 1
zebra(ws, STO_FIRST, STO_LAST, 8)
total_row(ws, r, ["סה״כ",
                  f"=SUM(B{STO_FIRST}:B{STO_LAST})",
                  f"=SUM(C{STO_FIRST}:C{STO_LAST})",
                  f"=SUM(D{STO_FIRST}:D{STO_LAST})",
                  f"=SUM(E{STO_FIRST}:E{STO_LAST})", "", "", ""],
          [None, FMT["shekel_whole"], FMT["shekel_whole"], FMT["int"], FMT["int"],
           None, None, None])
STO_TOTAL = r
add_table(ws, "טבלת_חנויות", f"A{STO_HDR}:H{STO_LAST}")
freeze_header(ws, STO_FIRST)
r += 2
note(ws, r, "סבב מלאי = כמה פעמים המלאי מתחלף בחודש. גבוה יותר = הסחורה זזה מהר יותר.")

STO = "'סיכום חנויות'!"
STO_NAME_RNG = f"{STO}$A${STO_FIRST}:$A${STO_LAST}"
STO_SALES_RNG = f"{STO}$B${STO_FIRST}:$B${STO_LAST}"

# ===========================================================================
# הכנסות
# ===========================================================================
ws = wb.create_sheet("הכנסות")
hide_gridlines(ws)
r = page_title(ws, 1, "הכנסות לפי חנות — 2026",
               "ינואר מגיע מגיליון המלאי. שאר החודשים = ינואר × מקדם העונתיות.")
r = header_row(ws, r, ["חנות"] + MONTHS + ["סה״כ שנתי"],
               widths=[24] + [13] * 12 + [16])
REV_FIRST = r
for store in STORES:
    vals = [store, f'=SUMIF({STORE_RNG},$A{r},{SALES_RNG})']
    for m in range(1, 12):
        vals.append(f"=$B{r}*INDEX({SEASON},{m + 1})")
    vals.append(f"=SUM(B{r}:M{r})")
    data_row(ws, r, vals, [None] + [FMT["shekel_whole"]] * 13)
    r += 1
REV_LAST = r - 1
zebra(ws, REV_FIRST, REV_LAST, 14)
REV_TOTAL = r
total_row(ws, r, ["סה״כ הכנסות"] +
          [f"=SUM({get_column_letter(c)}{REV_FIRST}:{get_column_letter(c)}{REV_LAST})"
           for c in range(2, 15)],
          [None] + [FMT["shekel_whole"]] * 13)
freeze_header(ws, REV_FIRST)

# ===========================================================================
# הוצאות
# ===========================================================================
ws = wb.create_sheet("הוצאות")
hide_gridlines(ws)
r = page_title(ws, 1, "הוצאות תפעוליות — 2026",
               "הוצאות קבועות. עולות בהדרגה לאורך השנה.")
r = header_row(ws, r, ["סעיף הוצאה"] + MONTHS + ["סה״כ שנתי"],
               widths=[24] + [13] * 12 + [16])
EXP_FIRST = r
drift = [1.00, 0.99, 1.02, 1.01, 1.04, 1.06, 1.05, 1.07, 1.06, 1.08, 1.10, 1.15]
for name, base in EXPENSES:
    vals = [name] + [round(base * d, -2) for d in drift]
    vals.append(f"=SUM(B{r}:M{r})")
    data_row(ws, r, vals, [None] + [FMT["shekel_whole"]] * 13)
    r += 1
EXP_LAST = r - 1
zebra(ws, EXP_FIRST, EXP_LAST, 14)
EXP_TOTAL = r
total_row(ws, r, ["סה״כ הוצאות"] +
          [f"=SUM({get_column_letter(c)}{EXP_FIRST}:{get_column_letter(c)}{EXP_LAST})"
           for c in range(2, 15)],
          [None] + [FMT["shekel_whole"]] * 13)
freeze_header(ws, EXP_FIRST)

# ===========================================================================
# רווח והפסד
# ===========================================================================
ws = wb.create_sheet("רווח והפסד")
hide_gridlines(ws)
r = page_title(ws, 1, "רווח והפסד — 2026",
               "מלמעלה למטה: כמה נכנס, מה ירד ממנו, ומה נשאר בסוף.")
r = header_row(ws, r, ["סעיף"] + MONTHS + ["סה״כ שנתי"],
               widths=[24] + [13] * 12 + [16])
PL_FIRST = r
PL = [
    ("הכנסות ברוטו",       lambda c, row: f"='הכנסות'!{c}{REV_TOTAL}", False),
    ("בניכוי מע״מ",        lambda c, row: f"=-{c}{PL_FIRST}*{VAT}/(1+{VAT})", False),
    ("הכנסות נטו",         lambda c, row: f"={c}{PL_FIRST}+{c}{PL_FIRST + 1}", True),
    ("בניכוי עלות הסחורה", lambda c, row: f"=-{c}{PL_FIRST + 2}*{COGS}", False),
    ("רווח גולמי",         lambda c, row: f"={c}{PL_FIRST + 2}+{c}{PL_FIRST + 3}", True),
    ("בניכוי הוצאות תפעול", lambda c, row: f"=-'הוצאות'!{c}{EXP_TOTAL}", False),
    ("רווח תפעולי",        lambda c, row: f"={c}{PL_FIRST + 4}+{c}{PL_FIRST + 5}", True),
]
for idx, (label, fn, is_sub) in enumerate(PL):
    row = PL_FIRST + idx
    vals = [label] + [fn(get_column_letter(2 + m), row) for m in range(12)]
    vals.append(f"=SUM(B{row}:M{row})")
    if is_sub:
        total_row(ws, row, vals, [None] + [FMT["shekel_neg"]] * 13)
    else:
        data_row(ws, row, vals, [None] + [FMT["shekel_neg"]] * 13)
GROSS_ROW = PL_FIRST + 4
OP_ROW = PL_FIRST + 6
NET_REV_ROW = PL_FIRST + 2
r = PL_FIRST + len(PL) + 1
r = section(ws, r, "שיעורי רווחיות", 14)
for label, num in (("שיעור רווח גולמי", GROSS_ROW), ("שיעור רווח תפעולי", OP_ROW)):
    vals = [label] + [f"=IFERROR({get_column_letter(2 + m)}{num}/"
                      f"{get_column_letter(2 + m)}{NET_REV_ROW},0)" for m in range(12)]
    vals.append(f"=IFERROR(N{num}/N{NET_REV_ROW},0)")
    data_row(ws, r, vals, [None] + [FMT["pct"]] * 13)
    r += 1
flag_negative(ws, f"B{OP_ROW}:N{OP_ROW}")
freeze_header(ws, PL_FIRST)

# ===========================================================================
# תזרים מזומנים
# ===========================================================================
ws = wb.create_sheet("תזרים מזומנים")
hide_gridlines(ws)
r = page_title(ws, 1, "תזרים מזומנים — 2026",
               "כמה מזומן נכנס ויצא בפועל בכל חודש, ומה נשאר בקופה.")
r = header_row(ws, r, ["חודש", "תאריך סוף חודש", "תקבולים", "תשלומים",
                       "תזרים נטו", "יתרה בקופה"],
               widths=[16, 20, 18, 18, 18, 20])
CF_FIRST = r
for m in range(12):
    row = CF_FIRST + m
    eom = datetime.date(2026, m + 1, 1)
    eom = (eom.replace(day=28) + datetime.timedelta(days=4)).replace(day=1) \
        - datetime.timedelta(days=1)
    col = get_column_letter(2 + m)
    data_row(ws, row, [
        MONTHS[m], eom,
        f"='הכנסות'!{col}{REV_TOTAL}",
        f"=-'הוצאות'!{col}{EXP_TOTAL}-'רווח והפסד'!{col}{NET_REV_ROW}*{COGS}",
        f"=C{row}+D{row}",
        (f"={OPENING_CASH}+E{row}" if m == 0 else f"=F{row - 1}+E{row}"),
    ], [None, FMT["date"], FMT["shekel_neg"], FMT["shekel_neg"],
        FMT["shekel_neg"], FMT["shekel_neg"]],
        align=["right", "center", "right", "right", "right", "right"])
CF_LAST = CF_FIRST + 11
zebra(ws, CF_FIRST, CF_LAST, 6)
flag_negative(ws, f"F{CF_FIRST}:F{CF_LAST}")
add_table(ws, "טבלת_תזרים", f"A{CF_FIRST - 1}:F{CF_LAST}")
freeze_header(ws, CF_FIRST)

# ===========================================================================
# לוח בקרה — built last, moved to the front
# ===========================================================================
ws = wb.create_sheet("לוח בקרה")
wb.move_sheet("לוח בקרה", offset=-(len(wb.sheetnames) - 1))
hide_gridlines(ws)
for c in range(1, 15):
    ws.column_dimensions[get_column_letter(c)].width = 15

r = page_title(ws, 1, "אופנת גלים בע״מ — לוח בקרה 2026",
               "נתוני דמה. כל המספרים מתעדכנים אוטומטית מהגיליונות האחרים.")

r = section(ws, r, "המספרים החשובים", 14)
r += 1
KPIS = [
    ("מכירות בחודש", f"='סיכום חנויות'!B{STO_TOTAL}", FMT["shekel_whole"], "כל 6 החנויות יחד"),
    ("שווי המלאי", f"='סיכום חנויות'!C{STO_TOTAL}", FMT["shekel_whole"], "לפי מחיר עלות"),
    ("פריטים בהתראה", f"='סיכום חנויות'!D{STO_TOTAL}+'סיכום חנויות'!E{STO_TOTAL}",
     FMT["int"], "אזלו או קריטיים"),
    ("רווח תפעולי שנתי", f"='רווח והפסד'!N{OP_ROW}", FMT["shekel_whole"], "אחרי כל ההוצאות"),
    ("שיעור רווח גולמי", f"='רווח והפסד'!N{GROSS_ROW}/'רווח והפסד'!N{NET_REV_ROW}",
     FMT["pct"], "מה נשאר אחרי עלות הסחורה"),
    ("מזומן בסוף השנה", f"='תזרים מזומנים'!F{CF_LAST}", FMT["shekel_whole"], "יתרה צפויה בקופה"),
]
for i, (label, formula, fmt, cap) in enumerate(KPIS):
    kpi(ws, r + (i // 3) * 4, 1 + (i % 3) * 3, label, formula, fmt, cap, width=3)
r += 8

r = section(ws, r, "מה חשוב לדעת", 14)
r += 1
INSIGHTS = [
    f'="1. החנות המובילה: "&INDEX({STO_NAME_RNG},MATCH(MAX({STO_SALES_RNG}),{STO_SALES_RNG},0))'
    f'&" — "&TEXT(MAX({STO_SALES_RNG}),"#,##0")&" ש""ח בחודש."',

    f'="2. החנות החלשה ביותר: "&INDEX({STO_NAME_RNG},MATCH(MIN({STO_SALES_RNG}),{STO_SALES_RNG},0))'
    f'&" — "&TEXT(MIN({STO_SALES_RNG}),"#,##0")&" ש""ח בחודש. פער של "'
    f'&TEXT(MAX({STO_SALES_RNG})/MIN({STO_SALES_RNG}),"0.0")&" מהחנות המובילה."',

    f'="3. הפריט שנגמר הכי מהר: "&INDEX({ITEM_RNG},MATCH(MIN({KEY_RNG}),{KEY_RNG},0))'
    f'&" מידה "&INDEX({SIZE_RNG},MATCH(MIN({KEY_RNG}),{KEY_RNG},0))'
    f'&" ב"&INDEX({STORE_RNG},MATCH(MIN({KEY_RNG}),{KEY_RNG},0))'
    f'&" — נשארו "&TEXT(MIN({COVER_RNG}),"0.0")&" שבועות מלאי."',

    f'="4. "&TEXT(COUNTIF({STATUS_RNG},"אזל")+COUNTIF({STATUS_RNG},"קריטי"),"0")'
    f'&" פריטים דורשים הזמנה מיידית, ו־"&TEXT(COUNTIF({STATUS_RNG},"נמוך"),"0")'
    f'&" פריטים מתקרבים לחוסר."',

    f'="5. "&TEXT(SUMIF({STATUS_RNG},"עודף",{VALUE_RNG}),"#,##0")&" ש""ח תקועים במלאי עודף — '
    f'"&TEXT(SUMIF({STATUS_RNG},"עודף",{VALUE_RNG})/SUM({VALUE_RNG}),"0%")&" מכלל המלאי."',

    f'="6. הרווח התפעולי הוא "&TEXT(\'רווח והפסד\'!N{OP_ROW}/\'רווח והפסד\'!N{NET_REV_ROW},"0.0%")'
    f'&" מההכנסות. "&IF(\'רווח והפסד\'!N{OP_ROW}<0,"החברה מפסידה.",'
    f'IF(\'רווח והפסד\'!N{OP_ROW}/\'רווח והפסד\'!N{NET_REV_ROW}<0.05,'
    f'"זה שולי — כל ירידה במכירות תמחק אותו.","זה בריא."))',
]
for line in INSIGHTS:
    insight(ws, r, line, 14)
    r += 1

r += 1
r = section(ws, r, "התראות מלאי — 10 הפריטים הדחופים ביותר", 9)
r = header_row(ws, r, ["חנות", "מק״ט", "פריט", "מידה", "מלאי (יח׳)",
                       "שבועות כיסוי", "סטטוס"])
ALERT_FIRST = r
# The 10 lowest-cover rows. MATCH runs against the unique sort key, not the raw cover
# column — matching on cover would return the same row for every tied value.
for k in range(1, 11):
    pos = f"MATCH(SMALL({KEY_RNG},{k}),{KEY_RNG},0)"
    data_row(ws, r, [
        f"=INDEX({STORE_RNG},{pos})",
        f"=INDEX({SKU_RNG},{pos})",
        f"=INDEX({ITEM_RNG},{pos})",
        f"=INDEX({SIZE_RNG},{pos})",
        f"=INDEX({STOCK_RNG},{pos})",
        f"=INDEX({COVER_RNG},{pos})",
        f"=INDEX({STATUS_RNG},{pos})",
    ], [None, None, None, None, FMT["int"], "0.0", None],
        align=["right", "center", "right", "center", "center", "center", "center"])
    r += 1
ALERT_LAST = r - 1
zebra(ws, ALERT_FIRST, ALERT_LAST, 7)
status_colours(ws, f"G{ALERT_FIRST}:G{ALERT_LAST}",
               bad=["אזל", "קריטי"], warn=["נמוך"], good=["תקין"], info=["עודף"])

# charts
sto_ws = wb["סיכום חנויות"]
bar = BarChart()
bar.type = "bar"
bar.title = "מכירות חודשיות לפי חנות"
bar.add_data(Reference(sto_ws, min_col=2, min_row=STO_HDR, max_row=STO_LAST),
             titles_from_data=True)
bar.set_categories(Reference(sto_ws, min_col=1, min_row=STO_FIRST, max_row=STO_LAST))
bar.dLbls = DataLabelList()
bar.dLbls.showVal = True
bar.legend = None
bar.height, bar.width = 9, 18
# Anchored past the KPI tiles (cols A–I). In an RTL sheet column K sits to the LEFT of A.
ws.add_chart(bar, "K4")

pl_ws = wb["רווח והפסד"]
line = LineChart()
line.title = "רווח תפעולי לפי חודש"
line.add_data(Reference(pl_ws, min_col=2, max_col=13, min_row=OP_ROW, max_row=OP_ROW),
              from_rows=True)
line.set_categories(Reference(pl_ws, min_col=2, max_col=13,
                              min_row=PL_FIRST - 1, max_row=PL_FIRST - 1))
line.legend = None
line.height, line.width = 9, 18
ws.add_chart(line, "K26")

# ===========================================================================
rtl_workbook(wb)
wb.save(OUT)
# מידה holds labels ("32", "L", "יחיד"), not quantities — Excel would flag the
# numeric-looking ones with a green triangle.
suppress_text_number_warning(OUT, {"מלאי לפי חנות": f"C{INV_FIRST}:C{INV_LAST}"})
print(f"saved: {OUT}")
print(f"sheets: {', '.join(wb.sheetnames)}")
print(f"inventory rows: {INV_LAST - INV_FIRST + 1} ({len(ITEMS)} items x {len(STORES)} stores)")
